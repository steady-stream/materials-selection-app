import argparse
import csv
import os
import re
from collections import Counter

import openpyxl

KEYWORDS = {
    "name",
    "product",
    "model",
    "modelnumber",
    "manufacturer",
    "vendor",
    "description",
    "category",
    "unit",
    "tier",
    "collection",
    "color",
    "finish",
    "price",
    "cost",
    "sku",
    "url",
    "image",
    "material",
    "item",
}

PRODUCT_HEADERS = {
    "name": ["name", "product", "productname", "item", "itemname"],
    "modelNumber": ["model", "modelnumber", "model#", "sku", "part", "partnumber"],
    "manufacturer": ["manufacturer", "brand", "mfg"],
    "vendor": ["vendor", "supplier", "store"],
    "description": ["description", "details", "spec"],
    "category": ["category", "section", "group", "type"],
    "unit": ["unit", "uom"],
    "price": ["price", "cost", "unitcost", "amount"],
    "url": ["url", "link", "producturl", "website"],
    "imageUrl": ["image", "imageurl", "photo"],
    "color": ["color"],
    "finish": ["finish"],
    "collection": ["collection", "series", "line"],
}

URL_RE = re.compile(r"https?://", re.IGNORECASE)
MODEL_RE = re.compile(r"[A-Za-z]{1,8}[-_/]?[A-Za-z0-9]{2,}")
PROJECT_TAB_RE = re.compile(r"\\b(\\d{4,6}|thc|lnc|cjc|lnx|thpt|bath|kitchen|basement|shower|condo)\\b", re.IGNORECASE)


def normalize(value):
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = re.sub(r"[^a-z0-9]+", "", text)
    return text


def looks_like_number(value):
    if value is None:
        return False
    text = str(value).strip().replace("$", "").replace(",", "")
    try:
        float(text)
        return True
    except Exception:
        return False


def header_score(row_values):
    score = 0
    hits = []
    for val in row_values:
        n = normalize(val)
        if not n:
            continue
        if n in KEYWORDS:
            score += 2
            hits.append(n)
        else:
            for kw in KEYWORDS:
                if kw in n:
                    score += 1
                    hits.append(kw)
                    break
    return score, sorted(set(hits))


def detect_header(rows):
    best_idx = None
    best_score = -1
    best_hits = []
    for idx, row in enumerate(rows[:25], start=1):
        score, hits = header_score(row)
        if score > best_score:
            best_score = score
            best_idx = idx
            best_hits = hits
    return best_idx, best_score, best_hits


def classify_sheet(name, header_hits, signal_score):
    n = name.lower().strip()
    if any(k in n for k in ["vendor", "manufacturer", "product", "service", "finish", "frequent"]):
        return "reference/master"
    if "template" in n:
        return "template"
    if PROJECT_TAB_RE.search(n):
        return "project"
    if signal_score >= 10 or len(header_hits) >= 3:
        return "likely_extractable"
    return "other"


def find_field_columns(header_row):
    field_to_col = {}
    normalized = [normalize(v) for v in header_row]
    for field, aliases in PRODUCT_HEADERS.items():
        for i, col in enumerate(normalized):
            if not col:
                continue
            if any(alias in col for alias in aliases):
                field_to_col[field] = i
                break
    return field_to_col


def row_nonempty_count(row):
    return sum(1 for cell in row if str(cell).strip() if cell is not None)


def cell_as_text(value):
    if value is None:
        return ""
    return str(value).strip()


def extract_candidate_rows(ws, sheet_name, header_idx, field_cols, max_rows=800):
    candidates = []
    if not header_idx:
        return candidates

    target_fields = ["name", "modelNumber", "manufacturer", "description", "category", "vendor", "price", "url", "imageUrl", "color", "finish", "collection"]

    for r_idx, row in enumerate(ws.iter_rows(min_row=header_idx + 1, max_row=min(ws.max_row, header_idx + max_rows), values_only=True), start=header_idx + 1):
        if row is None:
            continue

        nonempty = row_nonempty_count(row)
        if nonempty == 0:
            continue

        rec = {"sheet": sheet_name, "row": r_idx}
        for f in target_fields:
            col = field_cols.get(f)
            rec[f] = cell_as_text(row[col]) if col is not None and col < len(row) else ""

        name_like = bool(rec["name"]) or bool(rec["description"])
        model_like = bool(rec["modelNumber"]) or any(MODEL_RE.search(cell_as_text(c) or "") for c in row[: min(len(row), 12)])
        manufacturer_like = bool(rec["manufacturer"]) or any(str(c).strip().lower() in {"kohler", "moen", "delta", "panasonic", "daltile", "american standard"} for c in row if c)

        if name_like and (model_like or manufacturer_like):
            rec["raw_nonempty_cells"] = nonempty
            candidates.append(rec)

    return candidates


def main():
    parser = argparse.ArgumentParser(description="Analyze workbook tabs for product extraction potential")
    parser.add_argument("workbook", help="Path to xlsx/xlsm workbook")
    parser.add_argument("--outdir", default="analysis/excel", help="Output directory")
    args = parser.parse_args()

    workbook_path = args.workbook
    outdir = args.outdir
    os.makedirs(outdir, exist_ok=True)

    wb = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)

    sheet_inventory_path = os.path.join(outdir, "workbook_tab_inventory.csv")
    sheet_fieldmap_path = os.path.join(outdir, "workbook_tab_fieldmaps.csv")
    candidate_rows_path = os.path.join(outdir, "workbook_product_candidate_rows.csv")

    inventory_rows = []
    fieldmap_rows = []
    all_candidates = []

    for sname in wb.sheetnames:
        ws = wb[sname]

        sampled = []
        model_hits = 0
        url_hits = 0
        numeric_hits = 0
        nonempty_rows = 0

        for r_i, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 200), values_only=True), start=1):
            sampled.append(row)
            if any(cell_as_text(c) for c in row):
                nonempty_rows += 1
            for c in row:
                txt = cell_as_text(c)
                if not txt:
                    continue
                if URL_RE.search(txt):
                    url_hits += 1
                if MODEL_RE.search(txt):
                    model_hits += 1
                if looks_like_number(txt):
                    numeric_hits += 1

        header_idx, h_score, header_hits = detect_header(sampled)
        header_row = sampled[header_idx - 1] if header_idx and header_idx - 1 < len(sampled) else []
        field_cols = find_field_columns(header_row)

        name_score = 0
        low_name = sname.lower()
        if "product" in low_name:
            name_score += 4
        if "vendor" in low_name or "manufacturer" in low_name:
            name_score += 2
        if "material" in low_name or "finish" in low_name:
            name_score += 2

        signal_score = name_score + h_score + min(model_hits, 10) + min(url_hits, 6)
        classification = classify_sheet(sname, header_hits, signal_score)

        inventory_rows.append({
            "sheet": sname,
            "max_row": ws.max_row,
            "max_col": ws.max_column,
            "sample_nonempty_rows_first200": nonempty_rows,
            "detected_header_row": header_idx or "",
            "header_hit_count": len(header_hits),
            "header_hits": "|".join(header_hits),
            "model_like_hits_first200": model_hits,
            "url_hits_first200": url_hits,
            "numeric_hits_first200": numeric_hits,
            "mapped_field_count": len(field_cols),
            "mapped_fields": "|".join(sorted(field_cols.keys())),
            "product_extract_signal_score": signal_score,
            "classification": classification,
        })

        fieldmap_rows.append({
            "sheet": sname,
            "header_row": header_idx or "",
            "field_map": "|".join(f"{k}:{v+1}" for k, v in sorted(field_cols.items())),
        })

        if classification in {"reference/master", "likely_extractable"} or len(field_cols) >= 3:
            all_candidates.extend(extract_candidate_rows(ws, sname, header_idx, field_cols))

    inventory_rows.sort(key=lambda x: x["product_extract_signal_score"], reverse=True)

    with open(sheet_inventory_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=list(inventory_rows[0].keys()))
        writer.writeheader()
        writer.writerows(inventory_rows)

    with open(sheet_fieldmap_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=["sheet", "header_row", "field_map"])
        writer.writeheader()
        writer.writerows(fieldmap_rows)

    if all_candidates:
        with open(candidate_rows_path, "w", newline="", encoding="utf-8") as f:
            fields = ["sheet", "row", "name", "modelNumber", "manufacturer", "description", "category", "vendor", "price", "url", "imageUrl", "color", "finish", "collection", "raw_nonempty_cells"]
            writer = csv.DictWriter(f, fieldnames=fields)
            writer.writeheader()
            writer.writerows(all_candidates)
    else:
        with open(candidate_rows_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["sheet", "row", "name", "modelNumber", "manufacturer", "description", "category", "vendor", "price", "url", "imageUrl", "color", "finish", "collection", "raw_nonempty_cells"])

    print(f"Analyzed workbook: {workbook_path}")
    print(f"Tab count: {len(wb.sheetnames)}")
    print(f"Wrote: {sheet_inventory_path}")
    print(f"Wrote: {sheet_fieldmap_path}")
    print(f"Wrote: {candidate_rows_path}")
    print(f"Candidate row count: {len(all_candidates)}")


if __name__ == "__main__":
    main()
