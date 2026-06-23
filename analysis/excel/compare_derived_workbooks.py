import csv
import re
from pathlib import Path

import openpyxl

BASE_DIR = Path("g:/Projects/MegaPros/MaterialsSelectionApp/WebPrototype/analysis/excel")
BASELINE_CSV = BASE_DIR / "workbook_product_candidate_rows_filtered.csv"

WORKBOOKS = [
    Path(r"E:/SteadyStreamDevelopment/Projects/MegaPros Materials Selection/ChatGPT/Product_Master_Audit_Pass.xlsx"),
    Path(r"E:/SteadyStreamDevelopment/Projects/MegaPros Materials Selection/ChatGPT/Product_Master_Catalog_v3.xlsx"),
    Path(r"E:/SteadyStreamDevelopment/Projects/MegaPros Materials Selection/ChatGPT/Product_Master_Enhanced.xlsx"),
    Path(r"E:/SteadyStreamDevelopment/Projects/MegaPros Materials Selection/ChatGPT/Product_Master_Extract.xlsx"),
    Path(r"E:/SteadyStreamDevelopment/Projects/MegaPros Materials Selection/ChatGPT/Product_Master_Refined.xlsx"),
    Path(r"E:/SteadyStreamDevelopment/Projects/MegaPros Materials Selection/ChatGPT/Product_Master_Warehouse.xlsx"),
    Path(r"E:/SteadyStreamDevelopment/Projects/MegaPros Materials Selection/CoPilot/compiled_products_db.xlsx"),
    Path(r"E:/SteadyStreamDevelopment/Projects/MegaPros Materials Selection/CoPilot/normalized_products_db.xlsx"),
    Path(r"E:/SteadyStreamDevelopment/Projects/MegaPros Materials Selection/CoPilot/normalized_products_db_v2.xlsx"),
    Path(r"E:/SteadyStreamDevelopment/Projects/MegaPros Materials Selection/CoPilot/normalized_products_db_v2_cleaned_v3.xlsx"),
]

FIELD_ALIASES = {
    "name": ["name", "product", "item", "title"],
    "model": ["model", "modelnumber", "model#", "sku", "partnumber", "part"],
    "manufacturer": ["manufacturer", "brand", "mfg"],
    "vendor": ["vendor", "supplier", "store"],
    "description": ["description", "details", "spec"],
    "category": ["category", "type", "group", "section"],
    "price": ["price", "cost", "unitcost", "amount"],
    "url": ["url", "link", "producturl", "website"],
    "finish": ["finish"],
    "color": ["color"],
    "collection": ["collection", "series", "line"],
}

NOISE_PATTERNS = [
    r"^total",
    r"^contract",
    r"^additional",
    r"^deposit",
    r"amount to finance",
    r"^install",
    r"^invoice",
    r"^eta",
    r"^remove",
    r"^paint",
    r"^labor",
    r"^allowance",
]

MODEL_CLEAN_RE = re.compile(r"^(model\s*#?:?|sku\s*:?|part\s*#?:?)\s*", re.IGNORECASE)
MODEL_TOKEN_RE = re.compile(r"[A-Za-z]{1,10}[-_/]?[A-Za-z0-9]{2,}")


def norm(text):
    if text is None:
        return ""
    s = str(text).strip().lower()
    return re.sub(r"[^a-z0-9]+", "", s)


def txt(value):
    return "" if value is None else str(value).strip()


def looks_like_header(row):
    score = 0
    for c in row:
        n = norm(c)
        if not n:
            continue
        for aliases in FIELD_ALIASES.values():
            if any(a in n for a in aliases):
                score += 1
                break
    return score


def detect_header(rows):
    best_idx = None
    best_score = -1
    for i, row in enumerate(rows[:25], start=1):
        sc = looks_like_header(row)
        if sc > best_score:
            best_score = sc
            best_idx = i
    return best_idx, best_score


def map_fields(header_row):
    mapped = {}
    normalized = [norm(c) for c in header_row]
    for field, aliases in FIELD_ALIASES.items():
        for idx, col in enumerate(normalized):
            if any(a in col for a in aliases):
                mapped[field] = idx
                break
    return mapped


def clean_model(value):
    s = txt(value)
    s = MODEL_CLEAN_RE.sub("", s)
    s = s.replace("\u00a0", " ").strip()
    if not s:
        return ""
    m = MODEL_TOKEN_RE.search(s)
    return m.group(0).upper() if m else s.upper()


def is_noise(name, desc):
    text = f"{name} {desc}".strip().lower()
    if not text:
        return True
    return any(re.search(p, text) for p in NOISE_PATTERNS)


def load_baseline_keys():
    rows = list(csv.DictReader(BASELINE_CSV.open(encoding="utf-8")))
    keys = set()
    model_set = set()
    for r in rows:
        name = txt(r.get("name", "")).lower()
        model = clean_model(r.get("modelNumber", ""))
        if name or model:
            keys.add((name, model))
        if model:
            model_set.add(model)
    return keys, model_set


def analyze_workbook(path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)

    extracted = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        sampled = list(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 40), values_only=True))
        header_idx, header_score = detect_header(sampled)
        if not header_idx or header_score < 2:
            continue

        header_row = sampled[header_idx - 1]
        fmap = map_fields(header_row)
        if len(fmap) < 2:
            continue

        for r_idx, row in enumerate(ws.iter_rows(min_row=header_idx + 1, max_row=ws.max_row, values_only=True), start=header_idx + 1):
            name = txt(row[fmap["name"]]) if "name" in fmap and fmap["name"] < len(row) else ""
            desc = txt(row[fmap["description"]]) if "description" in fmap and fmap["description"] < len(row) else ""
            model = txt(row[fmap["model"]]) if "model" in fmap and fmap["model"] < len(row) else ""
            manufacturer = txt(row[fmap["manufacturer"]]) if "manufacturer" in fmap and fmap["manufacturer"] < len(row) else ""
            vendor = txt(row[fmap["vendor"]]) if "vendor" in fmap and fmap["vendor"] < len(row) else ""
            category = txt(row[fmap["category"]]) if "category" in fmap and fmap["category"] < len(row) else ""
            price = txt(row[fmap["price"]]) if "price" in fmap and fmap["price"] < len(row) else ""
            url = txt(row[fmap["url"]]) if "url" in fmap and fmap["url"] < len(row) else ""
            finish = txt(row[fmap["finish"]]) if "finish" in fmap and fmap["finish"] < len(row) else ""
            color = txt(row[fmap["color"]]) if "color" in fmap and fmap["color"] < len(row) else ""
            collection = txt(row[fmap["collection"]]) if "collection" in fmap and fmap["collection"] < len(row) else ""

            nameish = bool(name or desc)
            model_clean = clean_model(model)
            support = bool(model_clean or manufacturer or vendor)
            if not (nameish and support):
                continue

            extracted.append({
                "workbook": path.name,
                "sheet": sheet,
                "row": r_idx,
                "name": name,
                "modelNumber": model_clean,
                "manufacturer": manufacturer,
                "vendor": vendor,
                "description": desc,
                "category": category,
                "price": price,
                "url": url,
                "finish": finish,
                "color": color,
                "collection": collection,
                "is_noise": "1" if is_noise(name, desc) else "0",
            })

    return extracted


def main():
    baseline_keys, baseline_models = load_baseline_keys()

    all_rows = []
    summary = []

    for wb in WORKBOOKS:
        if not wb.exists():
            continue
        rows = analyze_workbook(wb)
        all_rows.extend(rows)

        total = len(rows)
        clean = [r for r in rows if r["is_noise"] == "0"]
        with_model = [r for r in clean if r["modelNumber"]]

        unique_models = sorted({r["modelNumber"] for r in with_model if r["modelNumber"]})
        keyset = {(r["name"].strip().lower(), r["modelNumber"]) for r in clean}
        new_vs_baseline = [k for k in keyset if k not in baseline_keys and (k[0] or k[1])]
        new_models_vs_baseline = [m for m in unique_models if m not in baseline_models]

        summary.append({
            "workbook": wb.name,
            "total_candidate_rows": total,
            "clean_rows": len(clean),
            "rows_with_model": len(with_model),
            "unique_models": len(unique_models),
            "new_name_model_pairs_vs_baseline": len(new_vs_baseline),
            "new_models_vs_baseline": len(new_models_vs_baseline),
        })

    summary.sort(key=lambda r: (r["new_models_vs_baseline"], r["rows_with_model"], r["clean_rows"]), reverse=True)

    out_summary = BASE_DIR / "derived_workbook_comparison_summary.csv"
    with out_summary.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=list(summary[0].keys()) if summary else ["workbook"])
        writer.writeheader()
        writer.writerows(summary)

    out_rows = BASE_DIR / "derived_workbook_candidate_rows.csv"
    with out_rows.open("w", newline="", encoding="utf-8") as f:
        fields = [
            "workbook",
            "sheet",
            "row",
            "name",
            "modelNumber",
            "manufacturer",
            "vendor",
            "description",
            "category",
            "price",
            "url",
            "finish",
            "color",
            "collection",
            "is_noise",
        ]
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        writer.writerows(all_rows)

    merged_clean = [r for r in all_rows if r["is_noise"] == "0"]
    merged_by_key = {}
    for r in merged_clean:
        key = (r["name"].strip().lower(), r["modelNumber"])
        if key not in merged_by_key:
            merged_by_key[key] = r

    out_new = BASE_DIR / "derived_rows_new_vs_baseline.csv"
    with out_new.open("w", newline="", encoding="utf-8") as f:
        fields = [
            "workbook",
            "sheet",
            "row",
            "name",
            "modelNumber",
            "manufacturer",
            "vendor",
            "description",
            "category",
            "price",
            "url",
            "finish",
            "color",
            "collection",
            "is_noise",
        ]
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for key, row in merged_by_key.items():
            if key not in baseline_keys and (key[0] or key[1]):
                writer.writerow(row)

    print(f"wrote: {out_summary}")
    print(f"wrote: {out_rows}")
    print(f"wrote: {out_new}")
    print(f"derived_candidate_rows={len(all_rows)}")


if __name__ == "__main__":
    main()
